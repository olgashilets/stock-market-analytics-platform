
from __future__ import annotations

import argparse
import hashlib
from datetime import datetime, time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ВАЖНО:
# Тикеры ниже — стабильные демонстрационные идентификаторы.
# Их можно будет позже заменить на реальные коды БВФБ, если они понадобятся.
INSTRUMENTS = {
    "СберБанк": {"ticker": "SberB-S0009", "currency_code": "BYN", "market_segment": "shares"},
    "Приорбанк": {"ticker": "Prior-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "МАПИД": {"ticker": "MAPID-S0003", "currency_code": "BYN", "market_segment": "shares"},
    "Трест-35": {"ticker": "Tr35-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "СветлогорскЗЖБиИК": {"ticker": "SZGBI-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "Минскпромстрой": {"ticker": "MPS-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "Брестгазоаппарат": {"ticker": "BGA-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "Керамин": {"ticker": "Keram-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "БЭРН": {"ticker": "BERN-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "Белинвестбанк": {"ticker": "BIB-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "БрестМК": {"ticker": "BRMK-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "МозырьНПЗ": {"ticker": "MNPZ-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "СРСУ-3": {"ticker": "SRSU3-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "СлуцкСахарКомб": {"ticker": "SSK-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "ЦУМ Минск": {"ticker": "CUMM-S0001", "currency_code": "BYN", "market_segment": "shares"},
    "Молочный Мир": {"ticker": "MolMir-S0001", "currency_code": "BYN", "market_segment": "shares"},
}

OUTPUT_COLUMNS = [
    "Тикер",
    "Краткое наименование эмитента",
    "Валюта ценообразования",
    "Цена, вал.обр. (мин.)",
    "Цена, вал.обр. (срвз.)",
    "",  # в реальных выгрузках 6-й столбец может приходить без понятного заголовка
    "Доходность (мин., %)",
    "Доходность (макс., %)",
    "Доходность (срвз., %)",
    "Оборот (в вал.ценообр.)",
    "Оборот (в шт.)",
    "Количество сделок",
    "Срок",
    "Время сделки",
]

COLUMN_WIDTHS = {
    1: 18,
    2: 42,
    3: 24,
    4: 22,
    5: 22,
    6: 16,
    7: 20,
    8: 20,
    9: 20,
    10: 24,
    11: 18,
    12: 20,
    13: 10,
    14: 22,
}


def stable_seed(*parts: str) -> int:
    raw = "|".join(parts).encode("utf-8")
    return int(hashlib.sha256(raw).hexdigest()[:16], 16) % (2**32)


def get_business_days(month_end: pd.Timestamp) -> list[pd.Timestamp]:
    month_end = pd.Timestamp(month_end)
    month_start = month_end.replace(day=1)
    return list(pd.bdate_range(month_start, month_end))


def build_profiles(df: pd.DataFrame) -> dict[str, dict[str, float]]:
    profiles: dict[str, dict[str, float]] = {}
    total_months = len(df)

    for issuer in df.columns[1:]:
        series = pd.to_numeric(df[issuer], errors="coerce").fillna(0.0)
        nonzero_ratio = float((series > 0).sum() / total_months)
        price_median = float(series[series > 0].median()) if (series > 0).any() else 0.0

        profile_rng = np.random.default_rng(stable_seed("profile", issuer))
        turnover_base = 1000 + 12000 * nonzero_ratio + 5000 * profile_rng.random()

        # Более дорогие бумаги получают чуть больший денежный оборот
        if price_median > 100:
            turnover_base *= 2.0
        elif price_median > 20:
            turnover_base *= 1.4
        elif price_median > 5:
            turnover_base *= 1.15

        profiles[issuer] = {
            "nonzero_ratio": nonzero_ratio,
            "price_median": price_median,
            "turnover_base": turnover_base,
            "profile_noise": float(profile_rng.random()),
        }

    return profiles


def split_value(total: float, parts_count: int, rng: np.random.Generator) -> np.ndarray:
    if parts_count == 1:
        return np.array([total], dtype=float)
    weights = rng.dirichlet(np.ones(parts_count))
    return total * weights


def generate_rows_for_month(row: pd.Series, profiles: dict[str, dict[str, float]]) -> list[dict]:
    month_end = pd.Timestamp(row["Дата"])
    business_days = get_business_days(month_end)
    generated_rows: list[dict] = []

    for issuer_name in row.index[1:]:
        monthly_price = pd.to_numeric(row[issuer_name], errors="coerce")

        if pd.isna(monthly_price) or float(monthly_price) <= 0:
            continue

        monthly_price = float(monthly_price)

        if issuer_name not in INSTRUMENTS:
            raise KeyError(
                f"Для инструмента '{issuer_name}' нет записи в словаре INSTRUMENTS."
            )

        profile = profiles[issuer_name]
        month_rng = np.random.default_rng(stable_seed(issuer_name, str(month_end.date())))

        # Чем чаще инструмент торгуется по истории, тем выше доля активных дней
        activity_ratio = 0.18 + 0.72 * profile["nonzero_ratio"]
        activity_ratio *= 0.90 + 0.20 * month_rng.random()
        activity_ratio = min(0.95, max(0.10, activity_ratio))

        active_days_count = max(1, min(len(business_days), round(len(business_days) * activity_ratio)))
        chosen_idx = sorted(month_rng.choice(len(business_days), size=active_days_count, replace=False))
        active_days = [business_days[i] for i in chosen_idx]

        volatility = 0.010 + 0.030 * (1 - profile["nonzero_ratio"]) + 0.010 * profile["profile_noise"]
        trend_strength = 0.008 + 0.020 * profile["profile_noise"]

        shocks = month_rng.normal(0, volatility, size=active_days_count)
        trend = np.linspace(-trend_strength, trend_strength, active_days_count)

        day_avg_prices = monthly_price * (1 + shocks + trend)
        day_avg_prices = np.clip(day_avg_prices, monthly_price * 0.50, None)
        # Подгоняем среднее значение активных дней к месячному ориентиру
        day_avg_prices *= monthly_price / day_avg_prices.mean()

        base_turnover = profile["turnover_base"]
        turnover_noise = 0.70 + 0.65 * month_rng.random(active_days_count)
        shock_multiplier = 1 + 2.5 * np.abs(shocks)

        day_turnover_values = base_turnover * turnover_noise * shock_multiplier
        day_turnover_qty = day_turnover_values / day_avg_prices

        # На некоторых днях делаем 2–3 строки на один тикер и дату,
        # чтобы сырой слой был ближе к реальным выгрузкам БВФБ.
        multirow_probs = np.array([0.72, 0.22, 0.06])

        instrument_meta = INSTRUMENTS[issuer_name]

        for trade_day, day_avg, day_turnover_val, day_qty in zip(
            active_days,
            day_avg_prices,
            day_turnover_values,
            day_turnover_qty,
        ):
            sessions_count = int(month_rng.choice([1, 2, 3], p=multirow_probs))
            value_parts = split_value(float(day_turnover_val), sessions_count, month_rng)
            qty_parts = split_value(float(day_qty), sessions_count, month_rng)

            for session_idx, (session_value, session_qty) in enumerate(zip(value_parts, qty_parts), start=1):
                local_rng = np.random.default_rng(
                    stable_seed(issuer_name, str(trade_day.date()), str(session_idx))
                )

                spread = 0.004 + 0.018 * local_rng.random()
                avg_shift = local_rng.normal(0, 0.006)

                price_avg = max(day_avg * (1 + avg_shift), day_avg * 0.55)
                price_min = max(price_avg * (1 - spread), 0.000001)

                # 6-й столбец оставляем нейтральным price_aux, как и в схеме raw-слоя
                price_aux = max(price_min, price_avg * (1 - spread * 0.35 + 0.01 * local_rng.random()))

                # Для акций в примерах БВФБ доходность обычно 0
                yield_min = 0.0
                yield_max = 0.0
                yield_avg = 0.0

                # Сделки грубо масштабируем от количества бумаг
                deals_count = int(
                    max(
                        1,
                        round(np.sqrt(max(session_qty, 1)) * (0.8 + 0.5 * local_rng.random()))
                    )
                )

                generated_rows.append(
                    {
                        "Тикер": instrument_meta["ticker"],
                        "Краткое наименование эмитента": issuer_name,
                        "Валюта ценообразования": instrument_meta["currency_code"],
                        "Цена, вал.обр. (мин.)": round(float(price_min), 6),
                        "Цена, вал.обр. (срвз.)": round(float(price_avg), 6),
                        "": round(float(price_aux), 6),
                        "Доходность (мин., %)": round(float(yield_min), 2),
                        "Доходность (макс., %)": round(float(yield_max), 2),
                        "Доходность (срвз., %)": round(float(yield_avg), 2),
                        "Оборот (в вал.ценообр.)": round(float(session_value), 2),
                        "Оборот (в шт.)": round(float(session_qty), 6),
                        "Количество сделок": deals_count,
                        "Срок": None,
                        "Время сделки": datetime.combine(trade_day.date(), time(0, 0, 0)),
                    }
                )

    return generated_rows


def style_worksheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    for col_idx, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # Денежные/числовые поля
        for cell in row[3:11]:
            if cell.column in (10,):
                cell.number_format = '0.00'
            else:
                cell.number_format = '0.000000'

        # Количество сделок
        row[11].number_format = '0'
        # Срок
        row[12].number_format = '0'
        # Дата/время
        row[13].number_format = 'yyyy-mm-dd hh:mm:ss'


def write_workbook(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MarketStockResult"

    ws.append(OUTPUT_COLUMNS)

    if rows:
        # Сортируем как в примерах: сначала более свежие даты
        rows_sorted = sorted(
            rows,
            key=lambda x: (x["Время сделки"], x["Тикер"]),
            reverse=True,
        )
        for row in rows_sorted:
            ws.append([row[col] for col in OUTPUT_COLUMNS])

    style_worksheet(ws)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Генерация демо-архивов БВФБ-подобного формата из monthly data.xlsx"
    )
    parser.add_argument("--input", required=True, help="Путь к исходному data.xlsx")
    parser.add_argument("--output", required=True, help="Папка для сгенерированных xlsx-архивов")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(input_path)

    if "Дата" not in df.columns:
        raise ValueError("В исходном файле не найден столбец 'Дата'.")

    missing = [col for col in df.columns[1:] if col not in INSTRUMENTS]
    if missing:
        raise ValueError(
            "Для некоторых столбцов нет настроек в INSTRUMENTS: "
            + ", ".join(map(str, missing))
        )

    profiles = build_profiles(df)

    manifest_rows: list[dict] = []

    for _, month_row in df.iterrows():
        month_end = pd.Timestamp(month_row["Дата"])
        rows = generate_rows_for_month(month_row, profiles)

        file_name = f"MarketStockResult_{month_end.strftime('%Y_%m')}.xlsx"
        output_path = output_dir / file_name

        write_workbook(rows, output_path)

        manifest_rows.append(
            {
                "archive_file": file_name,
                "period_month": month_end.strftime("%Y-%m"),
                "rows_count": len(rows),
                "trade_days_in_file": int(pd.Series([r["Время сделки"].date() for r in rows]).nunique()) if rows else 0,
            }
        )

    manifest_df = pd.DataFrame(manifest_rows)
    manifest_df.to_csv(output_dir / "manifest.csv", index=False, sep=";")

    print(f"Создано xlsx-файлов: {len(manifest_rows)}")
    print(f"Непустых файлов: {sum(1 for r in manifest_rows if r['rows_count'] > 0)}")
    print(f"Папка результата: {output_dir}")


if __name__ == "__main__":
    main()
